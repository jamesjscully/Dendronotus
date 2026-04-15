from __future__ import annotations

import csv
import math
from pathlib import Path
from typing import Iterable

import matplotlib.pyplot as plt
import numpy as np
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent
EXTRACTED = ROOT / "extracted"
OUT_PNG = ROOT / "akira_all_data_overview.png"


TXT_FILES = [
    "130618-03 3 traces.txt",
    "130618-03 Si1.txt",
    "130618-03 Si2.txt",
    "130618-03 Si3.txt",
    "Fig5A.txt",
    "JN2016_Fig5A_150413_01_4494.txt",
    "JN2016_Fig5B_130917_01_3351.txt",
    "JN2016_Fig5C_130917_01_3310.txt",
    "JN2016_Fig5E_150413_01_3856.txt",
    "JN2016_Fig5F_150413_01_3714.txt",
    "JN2016_Fig5G_150413_01_3460.txt",
]

XLSX_FILE = "JNsc2019_Fig2.xlsx"

TRACE_COLORS = ["#1f77b4", "#2ca02c", "#d62728", "#9467bd", "#8c564b", "#e377c2"]


def clean_label(value: object, fallback: str) -> str:
    if value is None:
        return fallback
    text = str(value).strip().strip('"').strip("'")
    return text if text else fallback


def first_line_has_header(path: Path) -> bool:
    first_line = path.read_text(encoding="utf-8", errors="ignore").splitlines()[0]
    return any(ch.isalpha() for ch in first_line)


def load_txt_dataset(path: Path) -> tuple[list[str], np.ndarray]:
    has_header = first_line_has_header(path)
    if has_header:
        with path.open("r", encoding="utf-8", errors="ignore", newline="") as handle:
            reader = csv.reader(handle, delimiter="\t")
            raw_header = next(reader)
        header = [clean_label(col, f"col{i+1}") for i, col in enumerate(raw_header)]
        data = np.genfromtxt(path, delimiter="\t", skip_header=1)
    else:
        data = np.genfromtxt(path, delimiter="\t")
        if data.ndim == 1:
            data = data.reshape(-1, 1)
        header = [f"col{i+1}" for i in range(data.shape[1])]

    if data.ndim == 1:
        data = data.reshape(-1, 1)
    return header, data


def downsample(x: np.ndarray, ys: Iterable[np.ndarray], max_points: int = 8000) -> tuple[np.ndarray, list[np.ndarray]]:
    n = len(x)
    if n <= max_points:
        return x, [np.asarray(y) for y in ys]
    stride = max(1, math.ceil(n / max_points))
    return x[::stride], [np.asarray(y)[::stride] for y in ys]


def plot_txt_panel(ax: plt.Axes, path: Path) -> None:
    header, data = load_txt_dataset(path)
    ncols = data.shape[1]
    title = path.stem

    if ncols >= 4 and all("Time" in header[i] for i in range(0, ncols, 2) if i < len(header)):
        pair_count = ncols // 2
        labels = [clean_label(header[2 * i + 1], f"trace {i + 1}") for i in range(pair_count)]
        for i in range(pair_count):
            x = data[:, 2 * i]
            y = data[:, 2 * i + 1]
            mask = np.isfinite(x) & np.isfinite(y)
            x = x[mask]
            y = y[mask]
            if len(x) == 0:
                continue
            x = x - x[0]
            x_ds, (y_ds,) = downsample(x, [y])
            ax.plot(x_ds, y_ds, lw=0.7, color=TRACE_COLORS[i % len(TRACE_COLORS)], label=labels[i])
    else:
        x = data[:, 0]
        x = x[np.isfinite(x)]
        for i in range(1, ncols):
            y = data[:, i]
            mask = np.isfinite(data[:, 0]) & np.isfinite(y)
            xx = data[mask, 0]
            yy = y[mask]
            if len(xx) == 0:
                continue
            xx = xx - xx[0]
            xx_ds, (yy_ds,) = downsample(xx, [yy])
            ax.plot(xx_ds, yy_ds, lw=0.7, color=TRACE_COLORS[(i - 1) % len(TRACE_COLORS)], label=clean_label(header[i], f"trace {i}"))

    ax.set_title(title, fontsize=9)
    ax.set_xlabel("Time (s)", fontsize=8)
    ax.set_ylabel("Value", fontsize=8)
    ax.tick_params(labelsize=7, length=2)
    ax.grid(True, alpha=0.2, linewidth=0.4)
    handles, labels = ax.get_legend_handles_labels()
    if handles:
        ax.legend(loc="upper right", fontsize=6, frameon=False, ncol=min(2, len(handles)))


def sheet_points(ws) -> list[list[object]]:
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append(list(row))
    return rows


def numeric_pair_arrays(rows: list[list[object]], x_idx: int, y_idx: int) -> tuple[np.ndarray, np.ndarray]:
    xs = []
    ys = []
    for row in rows[1:]:
        if x_idx >= len(row) or y_idx >= len(row):
            continue
        x = row[x_idx]
        y = row[y_idx]
        if x is None or y is None:
            continue
        try:
            xs.append(float(x))
            ys.append(float(y))
        except (TypeError, ValueError):
            continue
    return np.asarray(xs, dtype=float), np.asarray(ys, dtype=float)


def plot_sheet_panel(ax: plt.Axes, xlsx_path: Path, sheet_name: str) -> None:
    wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    ws = wb[sheet_name]
    rows = sheet_points(ws)
    header = rows[0]
    ncols = max(len(r) for r in rows if r)

    pairs: list[tuple[int, int]] = []
    if sheet_name == "Fig 2Aiii":
        pairs = [(0, 1), (3, 4)]
    elif ncols >= 4:
        pairs = [(0, 1), (2, 3)]
    elif ncols >= 2:
        pairs = [(0, 1)]

    for i, (x_idx, y_idx) in enumerate(pairs):
        x, y = numeric_pair_arrays(rows, x_idx, y_idx)
        if len(x) == 0:
            continue
        label_y = clean_label(header[y_idx], f"series {i + 1}")
        color = TRACE_COLORS[i % len(TRACE_COLORS)]
        ax.scatter(x, y, s=10, alpha=0.8, color=color, edgecolors="none", label=label_y)

    ax.set_title(f"{xlsx_path.stem}: {sheet_name}", fontsize=9)
    ax.set_xlabel(clean_label(header[pairs[0][0]], "x"), fontsize=8)
    ax.set_ylabel("Value", fontsize=8)
    ax.tick_params(labelsize=7, length=2)
    ax.grid(True, alpha=0.2, linewidth=0.4)
    handles, labels = ax.get_legend_handles_labels()
    if handles:
        ax.legend(loc="best", fontsize=6, frameon=False)


def build_figure() -> Path:
    txt_paths = [EXTRACTED / name for name in TXT_FILES]
    xlsx_path = EXTRACTED / XLSX_FILE

    wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    sheet_names = wb.sheetnames
    panel_total = len(txt_paths) + len(sheet_names)
    ncols = 3
    nrows = math.ceil(panel_total / ncols)

    fig, axes = plt.subplots(nrows=nrows, ncols=ncols, figsize=(18, 3.6 * nrows), constrained_layout=True)
    axes = np.atleast_1d(axes).ravel()

    for ax, path in zip(axes, txt_paths):
        plot_txt_panel(ax, path)

    offset = len(txt_paths)
    for ax, sheet_name in zip(axes[offset:], sheet_names):
        plot_sheet_panel(ax, xlsx_path, sheet_name)

    for ax in axes[panel_total:]:
        ax.axis("off")

    fig.suptitle("Akira Bundle Overview: numeric traces and summary datasets", fontsize=14, y=1.01)
    fig.savefig(OUT_PNG, dpi=220, bbox_inches="tight")
    plt.close(fig)
    return OUT_PNG


if __name__ == "__main__":
    out = build_figure()
    print(out)
